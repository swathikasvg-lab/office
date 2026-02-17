# reports/desktop/excel_1003.py
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from datetime import datetime, timezone
import os
import math

# ------------------------------------------------------------
# Helpers (same spirit as pdf_1003.py)
# ------------------------------------------------------------
def _safe_get_series_value(series, field_name=None):
    """
    Return latest field value for a given Influx-style series dict:
    {
        "columns": [...],
        "values": [[time, val, ...], ...],
        "tags": {...}
    }
    If field_name is None or missing, falls back to common numeric fields.
    """
    if not series:
        return None
    cols = series.get("columns", [])
    vals = series.get("values", [])
    if not vals:
        return None

    # explicit field name
    if field_name and field_name in cols:
        idx = cols.index(field_name)
        return vals[-1][idx] if idx < len(vals[-1]) else None

    # common field names
    for try_name in (
        "value", "last", "used_percent", "download_mbps", "upload_mbps",
        "response_time_ms", "packet_loss_percent", "urlresponse", "mean"
    ):
        if try_name in cols:
            idx = cols.index(try_name)
            return vals[-1][idx] if idx < len(vals[-1]) else None

    # fallback to 2nd column
    if len(vals[-1]) >= 2:
        return vals[-1][1]
    return None


def _series_values_for_trend(series):
    """
    Prepare (times, values) for trend sheet.
    Tries urlresponse/mean/value/second column.
    """
    xs = []
    ys = []
    if not series:
        return xs, ys

    cols = series.get("columns", [])
    vals = series.get("values", [])

    # time index
    try:
        t_idx = cols.index("time")
    except ValueError:
        t_idx = 0

    # y index
    y_idx = None
    for name in ("urlresponse", "mean", "value"):
        if name in cols:
            y_idx = cols.index(name)
            break
    if y_idx is None:
        y_idx = 1 if len(cols) > 1 else None

    for row in vals:
        if t_idx < len(row):
            t_raw = row[t_idx]
        else:
            t_raw = None

        if y_idx is None or y_idx >= len(row):
            v_raw = None
        else:
            v_raw = row[y_idx]

        # Parse time into datetime or keep string
        xs.append(_try_parse_time(t_raw))

        # numeric value
        try:
            v_num = float(v_raw) if v_raw is not None else None
            if math.isfinite(v_num):
                ys.append(v_num)
            else:
                ys.append(None)
        except Exception:
            ys.append(None)

    return xs, ys


def _try_parse_time(t):
    """Try to parse Influx time formats into datetime (used for labels)."""
    if t is None:
        return None
    if isinstance(t, datetime):
        return t
    if isinstance(t, (int, float)):
        # guess seconds vs ms vs ns
        if t > 1e12:
            # ns -> s
            return datetime.fromtimestamp(t / 1e9, tz=timezone.utc)
        elif t > 1e9:
            # ms
            return datetime.fromtimestamp(t / 1000, tz=timezone.utc)
        else:
            return datetime.fromtimestamp(t, tz=timezone.utc)
    if isinstance(t, str):
        s = t.strip()
        try:
            if s.endswith("Z"):
                s2 = s[:-1] + "+00:00"
            else:
                s2 = s
            return datetime.fromisoformat(s2)
        except Exception:
            # maybe numeric-in-string
            try:
                num = float(s)
                return _try_parse_time(num)
            except Exception:
                return s
    return t


def _format_value(v, precision=1):
    if v is None:
        return "-"
    try:
        if isinstance(v, bool):
            return str(v)
        f = float(v)
        return f"{f:.{precision}f}"
    except Exception:
        return str(v)


# ------------------------------------------------------------
# Threshold coloring helpers
# ------------------------------------------------------------
RED_FILL    = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
GREEN_FILL  = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
HEADER_FILL = PatternFill(start_color="FFEAF3FF", end_color="FFEAF3FF", fill_type="solid")
CARD_HEADER_FILL = PatternFill(start_color="FFF5F9FF", end_color="FFF5F9FF", fill_type="solid")
BORDER_THIN = Border(
    left=Side(style="thin", color="FFD9D9D9"),
    right=Side(style="thin", color="FFD9D9D9"),
    top=Side(style="thin", color="FFD9D9D9"),
    bottom=Side(style="thin", color="FFD9D9D9"),
)


def apply_threshold_cpu_mem(cell, value):
    """
    CPU & Memory -> 70 (orange), 90 (red)
    """
    try:
        v = float(value)
    except Exception:
        return
    if v >= 90:
        cell.fill = RED_FILL
    elif v >= 70:
        cell.fill = ORANGE_FILL


def apply_threshold_speed(cell, value):
    """
    Download & Upload Mbps -> <10 (orange), <5 (red)
    Lower is worse.
    """
    try:
        v = float(value)
    except Exception:
        return
    if v < 5:
        cell.fill = RED_FILL
    elif v < 10:
        cell.fill = ORANGE_FILL


def apply_threshold_url_response(cell, value):
    """
    URL Response ms -> >100 (orange), >500 (red)
    """
    try:
        v = float(value)
    except Exception:
        return
    if v > 500:
        cell.fill = RED_FILL
    elif v > 100:
        cell.fill = ORANGE_FILL


def apply_threshold_packet_loss(cell, value):
    """
    Packet Loss % -> >30 (orange), >70 (red)
    """
    try:
        v = float(value)
    except Exception:
        return
    if v > 70:
        cell.fill = RED_FILL
    elif v > 30:
        cell.fill = ORANGE_FILL


def apply_threshold_latency(cell, value):
    """
    Latency ms -> >100 (orange), >250 (red)
    """
    try:
        v = float(value)
    except Exception:
        return
    if v > 250:
        cell.fill = RED_FILL
    elif v > 100:
        cell.fill = ORANGE_FILL


# ------------------------------------------------------------
# Main Excel builder
# ------------------------------------------------------------
def build_excel_1003(host, start, end, basic, net, updates, urlinfo_list, trend_series, mtr_data, customer):
    """
    Build an enterprise-grade Excel for Desktop Performance Report 1003.
    Mirrors the sections of pdf_1003.py but as a multi-sheet workbook.

    Returns: full path to generated .xlsx file
    """
    output_dir = "/usr/local/autointelli/opsduty-server/generated_reports"
    os.makedirs(output_dir, exist_ok=True)

    safe_host = str(host).replace("/", "_").replace(" ", "_")
    outfile = os.path.join(output_dir, f"Desktop_Performance_{safe_host}.xlsx")

    try:
        if os.path.exists(outfile):
            os.remove(outfile)
    except Exception:
        pass

    wb = Workbook()

    # Fonts
    title_font   = Font(bold=True, size=14, color="FF0B61D6")
    header_font  = Font(bold=True, size=11, color="FF0B61D6")
    table_hdr_ft = Font(bold=True)
    normal_font  = Font(size=10)

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left", vertical="center")

    # ============================================================
    # Extract core values from series (once)
    # ============================================================
    os_info = basic.get("os_info")
    cpu_ser = basic.get("cpu")
    mem_ser = basic.get("mem")
    disk_series = basic.get("disk", [])

    # OS Name
    os_name = "-"
    if os_info and os_info.get("values"):
        cols = os_info.get("columns", [])
        vals = os_info.get("values", [])
        if "os_name_1" in cols:
            os_name = vals[0][cols.index("os_name_1")]
        else:
            os_name = vals[0][1] if len(vals[0]) > 1 else "-"

    cpu_val = _safe_get_series_value(cpu_ser)
    mem_val = _safe_get_series_value(mem_ser, "used_percent")

    # Disk details
    disk_list = []
    for s in (disk_series or []):
        tags = s.get("tags", {})
        path = tags.get("path") or tags.get("mountpoint") or tags.get("device") or "unknown"
        vals = s.get("values", [])
        last_val = vals[-1][1] if vals else None
        disk_list.append((path, last_val))

    # Network
    speed = net.get("speed")
    gw    = net.get("gateway")
    download = _safe_get_series_value(speed, "download_mbps") or _safe_get_series_value(speed, "download")
    upload   = _safe_get_series_value(speed, "upload_mbps") or _safe_get_series_value(speed, "upload")
    loss     = _safe_get_series_value(gw, "packet_loss_percent") or _safe_get_series_value(gw, "packet_loss")
    rtime    = _safe_get_series_value(gw, "response_time_ms") or _safe_get_series_value(gw, "response_time")

    # Updates
    pending = _safe_get_series_value(updates.get("pending"), "last") or \
              _safe_get_series_value(updates.get("pending"), "pending_updates")
    upto    = _safe_get_series_value(updates.get("up_to_date"), "last") or \
              _safe_get_series_value(updates.get("up_to_date"), "is_up_to_date")

    # URL info list -> normalized rows
    url_rows = []
    for u in (urlinfo_list or []):
        target = u.get("target") or u.get("tags", {}).get("target") or "-"
        resp   = _safe_get_series_value(u, "urlresponse") or u.get("response") or u.get("Response Time") or u.get("value")
        stat   = u.get("status") or u.get("Status Code") or "-"
        url_rows.append((target, resp, stat))

    # Trend series
    trend_times, trend_vals = _series_values_for_trend(trend_series) if trend_series else ([], [])

    # ============================================================
    # SHEET 1: OVERVIEW
    # ============================================================
    ws_ov = wb.active
    ws_ov.title = "Overview"

    ws_ov["A1"] = "Desktop Performance Report"
    ws_ov["A1"].font = title_font

    ws_ov["A3"] = "Customer"
    ws_ov["B3"] = customer or "-"
    ws_ov["A4"] = "Host"
    ws_ov["B4"] = host
    ws_ov["A5"] = "Period"
    ws_ov["B5"] = f"{start} — {end}"

    for cell in ("A3", "A4", "A5"):
        ws_ov[cell].font = header_font

    # KPIs block
    ws_ov["A7"] = "Key Metrics"
    ws_ov["A7"].font = header_font

    headers = ["Metric", "Value"]
    ws_ov.append(headers)
    row_hdr = ws_ov[8]
    for cell in row_hdr:
        cell.font = table_hdr_ft
        cell.fill = HEADER_FILL
        cell.border = BORDER_THIN
        cell.alignment = center

    kpi_rows = [
        ("CPU Usage (%)", _format_value(cpu_val)),
        ("Memory Usage (%)", _format_value(mem_val)),
        ("Download Mbps", _format_value(download)),
        ("Upload Mbps", _format_value(upload)),
        ("Pending Updates", _format_value(pending, precision=0)),
    ]

    for name, val in kpi_rows:
        row = [name, val]
        ws_ov.append(row)

    # apply borders & threshold colors
    for r in range(9, 9 + len(kpi_rows)):
        ws_ov[f"A{r}"].border = BORDER_THIN
        ws_ov[f"B{r}"].border = BORDER_THIN
        ws_ov[f"A{r}"].alignment = left
        ws_ov[f"B{r}"].alignment = center

        metric = ws_ov[f"A{r}"].value
        value  = ws_ov[f"B{r}"].value

        if metric.startswith("CPU"):
            apply_threshold_cpu_mem(ws_ov[f"B{r}"], value)
        elif metric.startswith("Memory"):
            apply_threshold_cpu_mem(ws_ov[f"B{r}"], value)
        elif metric.startswith("Download"):
            apply_threshold_speed(ws_ov[f"B{r}"], value)
        elif metric.startswith("Upload"):
            apply_threshold_speed(ws_ov[f"B{r}"], value)

    ws_ov.column_dimensions["A"].width = 24
    ws_ov.column_dimensions["B"].width = 18

    # ============================================================
    # SHEET 2: BASIC INFO
    # ============================================================
    ws_b = wb.create_sheet("Basic Info")
    ws_b["A1"] = "Basic System Information"
    ws_b["A1"].font = title_font

    ws_b.append(["Field", "Value"])
    hdr = ws_b[2]
    for c in hdr:
        c.font = table_hdr_ft
        c.fill = HEADER_FILL
        c.border = BORDER_THIN
        c.alignment = center

    rows = [
        ("Hostname", host),
        ("Operating System", os_name),
        ("CPU Usage (%)", _format_value(cpu_val)),
        ("Memory Usage (%)", _format_value(mem_val)),
    ]
    ws_row = 3
    for label, val in rows:
        ws_b.append([label, val])
        ws_b[f"A{ws_row}"].border = BORDER_THIN
        ws_b[f"B{ws_row}"].border = BORDER_THIN
        ws_b[f"A{ws_row}"].alignment = left
        ws_b[f"B{ws_row}"].alignment = center
        if label == "CPU Usage (%)":
            apply_threshold_cpu_mem(ws_b[f"B{ws_row}"], val)
        if label == "Memory Usage (%)":
            apply_threshold_cpu_mem(ws_b[f"B{ws_row}"], val)
        ws_row += 1

    # Disk table
    ws_b.append([])
    ws_row += 1
    ws_b[f"A{ws_row}"] = "Disk Usage Details"
    ws_b[f"A{ws_row}"].font = header_font
    ws_row += 1

    ws_b.append(["Path", "Used (%)"])
    hdr2 = ws_b[ws_row]
    for c in hdr2:
        c.font = table_hdr_ft
        c.fill = HEADER_FILL
        c.border = BORDER_THIN
        c.alignment = center

    for path, val in disk_list:
        ws_row += 1
        ws_b.append([path, _format_value(val)])
        ws_b[f"A{ws_row}"].border = BORDER_THIN
        ws_b[f"B{ws_row}"].border = BORDER_THIN
        ws_b[f"A{ws_row}"].alignment = left
        ws_b[f"B{ws_row}"].alignment = center

    ws_b.column_dimensions["A"].width = 28
    ws_b.column_dimensions["B"].width = 18

    # ============================================================
    # SHEET 3: NETWORK
    # ============================================================
    ws_n = wb.create_sheet("Network")
    ws_n["A1"] = "Network Information"
    ws_n["A1"].font = title_font

    ws_n.append(["Metric", "Value"])
    hdr = ws_n[2]
    for c in hdr:
        c.font = table_hdr_ft
        c.fill = HEADER_FILL
        c.border = BORDER_THIN
        c.alignment = center

    net_rows = [
        ("Download Mbps", _format_value(download)),
        ("Upload Mbps", _format_value(upload)),
        ("Gateway Packet Loss (%)", _format_value(loss)),
        ("Gateway Response Time (ms)", _format_value(rtime)),
    ]

    row_idx = 3
    for label, val in net_rows:
        ws_n.append([label, val])
        ws_n[f"A{row_idx}"].border = BORDER_THIN
        ws_n[f"B{row_idx}"].border = BORDER_THIN
        ws_n[f"A{row_idx}"].alignment = left
        ws_n[f"B{row_idx}"].alignment = center

        if label.startswith("Download"):
            apply_threshold_speed(ws_n[f"B{row_idx}"], val)
        elif label.startswith("Upload"):
            apply_threshold_speed(ws_n[f"B{row_idx}"], val)
        elif label.startswith("Gateway Packet Loss"):
            apply_threshold_packet_loss(ws_n[f"B{row_idx}"], val)
        elif label.startswith("Gateway Response Time"):
            apply_threshold_latency(ws_n[f"B{row_idx}"], val)

        row_idx += 1

    ws_n.column_dimensions["A"].width = 30
    ws_n.column_dimensions["B"].width = 18

    # ============================================================
    # SHEET 4: UPDATES
    # ============================================================
    ws_u = wb.create_sheet("Updates")
    ws_u["A1"] = "System Update Status"
    ws_u["A1"].font = title_font

    ws_u.append(["Metric", "Value"])
    hdr = ws_u[2]
    for c in hdr:
        c.font = table_hdr_ft
        c.fill = HEADER_FILL
        c.border = BORDER_THIN
        c.alignment = center

    upd_rows = [
        ("Is Up-to-Date", str(bool(upto)) if upto is not None else "-"),
        ("Pending Updates", _format_value(pending, precision=0)),
    ]
    r = 3
    for label, val in upd_rows:
        ws_u.append([label, val])
        ws_u[f"A{r}"].border = BORDER_THIN
        ws_u[f"B{r}"].border = BORDER_THIN
        ws_u[f"A{r}"].alignment = left
        ws_u[f"B{r}"].alignment = center
        r += 1

    ws_u.column_dimensions["A"].width = 26
    ws_u.column_dimensions["B"].width = 18

    # ============================================================
    # SHEET 5: CRITICAL URLS
    # ============================================================
    ws_url = wb.create_sheet("Critical URLs")
    ws_url["A1"] = "Critical URL Reachability"
    ws_url["A1"].font = title_font

    ws_url.append(["URL", "Avg Response (ms)", "Status Code"])
    hdr = ws_url[2]
    for c in hdr:
        c.font = table_hdr_ft
        c.fill = HEADER_FILL
        c.border = BORDER_THIN
        c.alignment = center

    row_idx = 3
    if not url_rows:
        url_rows.append(("-", None, "-"))

    for target, resp, stat in url_rows:
        resp_str = _format_value(resp, precision=0)
        ws_url.append([target, resp_str, stat])
        ws_url[f"A{row_idx}"].border = BORDER_THIN
        ws_url[f"B{row_idx}"].border = BORDER_THIN
        ws_url[f"C{row_idx}"].border = BORDER_THIN
        ws_url[f"A{row_idx}"].alignment = left
        ws_url[f"B{row_idx}"].alignment = center
        ws_url[f"C{row_idx}"].alignment = center

        apply_threshold_url_response(ws_url[f"B{row_idx}"], resp_str)

        row_idx += 1

    ws_url.column_dimensions["A"].width = 50
    ws_url.column_dimensions["B"].width = 20
    ws_url.column_dimensions["C"].width = 14

    # ============================================================
    # SHEET 6: TREND
    # ============================================================
    ws_t = wb.create_sheet("Trend")
    ws_t["A1"] = "Response Time Trend (Sample URL)"
    ws_t["A1"].font = title_font

    ws_t.append(["Time", "Response (ms)"])
    hdr = ws_t[2]
    for c in hdr:
        c.font = table_hdr_ft
        c.fill = HEADER_FILL
        c.border = BORDER_THIN
        c.alignment = center

    row_idx = 3
    for t_val, v in zip(trend_times, trend_vals):
        if isinstance(t_val, datetime):
            label = t_val.strftime("%Y-%m-%d %H:%M")
        else:
            label = str(t_val)
        ws_t.append([label, v])
        ws_t[f"A{row_idx}"].border = BORDER_THIN
        ws_t[f"B{row_idx}"].border = BORDER_THIN
        ws_t[f"A{row_idx}"].alignment = left
        ws_t[f"B{row_idx}"].alignment = center
        row_idx += 1

    ws_t.column_dimensions["A"].width = 26
    ws_t.column_dimensions["B"].width = 16

    # Build line chart if we have enough points
    if row_idx > 3:
        chart = LineChart()
        chart.title = "URL Response Time"
        chart.y_axis.title = "ms"
        chart.x_axis.title = "Time"

        data_ref = Reference(ws_t, min_col=2, min_row=2, max_row=row_idx - 1)
        cats_ref = Reference(ws_t, min_col=1, min_row=3, max_row=row_idx - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 10
        chart.width = 24

        ws_t.add_chart(chart, "D2")

    # ============================================================
    # SHEET 7: TRACE ROUTE
    # ============================================================
    ws_tr = wb.create_sheet("Trace Route")
    ws_tr["A1"] = "Trace Route Information"
    ws_tr["A1"].font = title_font

    start_row = 3
    for target, item in (mtr_data or {}).items():
        ws_tr.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
        cell = ws_tr.cell(row=start_row, column=1, value=f"Target: {target}")
        cell.font = header_font
        cell.alignment = left
        start_row += 1

        # Packet Loss
        pl_series_list = item.get("packet_loss", []) or []
        pl_data = []
        for s in pl_series_list:
            tags = s.get("tags", {})
            hop = tags.get("hop", "-")
            ip  = tags.get("ip", "-")
            vals = s.get("values", [])
            mean_val = vals[-1][1] if vals else None
            try:
                hop_num = int(str(hop).strip())
            except Exception:
                hop_num = 99999
            pl_data.append((hop_num, ip, mean_val))

        ws_tr.append(["Hop", "IP", "Packet Loss (%)"])
        hdr = ws_tr[start_row]
        for c in hdr:
            c.font = table_hdr_ft
            c.fill = HEADER_FILL
            c.border = BORDER_THIN
            c.alignment = center

        row = start_row + 1
        if not pl_data:
            ws_tr.append(["-", "-", "-"])
            for col in range(1, 4):
                cell = ws_tr.cell(row=row, column=col)
                cell.border = BORDER_THIN
                cell.alignment = center
            row += 1
        else:
            pl_data.sort(key=lambda x: x[0])
            for hop_num, ip, val in pl_data:
                ws_tr.append([hop_num, ip, _format_value(val)])
                for col in range(1, 4):
                    cell = ws_tr.cell(row=row, column=col)
                    cell.border = BORDER_THIN
                    cell.alignment = center if col != 2 else left
                # threshold coloring on packet loss
                apply_threshold_packet_loss(ws_tr.cell(row=row, column=3), val)
                row += 1

        start_row = row + 1

        # Latency
        ws_tr.append(["Hop", "IP", "Latency (ms)"])
        hdr = ws_tr[start_row]
        for c in hdr:
            c.font = table_hdr_ft
            c.fill = HEADER_FILL
            c.border = BORDER_THIN
            c.alignment = center

        lat_series_list = item.get("latency", []) or []
        lat_data = []
        for s in lat_series_list:
            tags = s.get("tags", {})
            hop = tags.get("hop", "-")
            ip  = tags.get("ip", "-")
            vals = s.get("values", [])
            last_val = vals[-1][1] if vals else None
            try:
                hop_num = int(str(hop).strip())
            except Exception:
                hop_num = 99999
            lat_data.append((hop_num, ip, last_val))

        row = start_row + 1
        if not lat_data:
            ws_tr.append(["-", "-", "-"])
            for col in range(1, 4):
                cell = ws_tr.cell(row=row, column=col)
                cell.border = BORDER_THIN
                cell.alignment = center
            row += 1
        else:
            lat_data.sort(key=lambda x: x[0])
            for hop_num, ip, val in lat_data:
                ws_tr.append([hop_num, ip, _format_value(val)])
                for col in range(1, 4):
                    cell = ws_tr.cell(row=row, column=col)
                    cell.border = BORDER_THIN
                    cell.alignment = center if col != 2 else left
                apply_threshold_latency(ws_tr.cell(row=row, column=3), val)
                row += 1

        start_row = row + 2  # spacing before next target

    ws_tr.column_dimensions["A"].width = 8
    ws_tr.column_dimensions["B"].width = 28
    ws_tr.column_dimensions["C"].width = 18
    ws_tr.column_dimensions["D"].width = 4

    # Save workbook
    wb.save(outfile)
    return outfile

