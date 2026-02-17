import os
import tempfile
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

from .pdf_1007 import IST, _query_port_timeseries


def build_excel(targets, start, end, interval):
    """
    Build Excel Port Performance Report (1007).

    targets: list of dicts: {"server": str, "port": int}
    start, end: UTC ISO strings (e.g. "2025-12-02T06:00:00Z")
    interval: InfluxDB GROUP BY interval (e.g. "1m", "5m")
    """
    out_dir = tempfile.gettempdir()
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    outfile = os.path.join(out_dir, f"rpt_1007_port_{ts}.xlsx")

    wb = openpyxl.Workbook()

    # Basic styles
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_header = PatternFill("solid", fgColor="DCE6F7")
    fill_altrow = PatternFill("solid", fgColor="F7FAFF")

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Convert report range to IST for display
    start_ist = datetime.fromisoformat(start.replace("Z", "+00:00")).astimezone(IST)
    end_ist = datetime.fromisoformat(end.replace("Z", "+00:00")).astimezone(IST)

    headers = [
        "Time (IST)",
        "Mean Resp (ms)",
        "Max Resp (ms)",
        "Attempts",
        "Failures",
        "Availability (%)",
    ]

    for idx, t in enumerate(targets):
        server = t["server"]
        port = t["port"]
        if port is None:
            continue

        # Sheet name – limited to 31 chars
        sheet_name = f"{server}_{port}"[:31] or f"target_{idx+1}"
        ws = wb.create_sheet(title=sheet_name)

        # ----------------- SHEET HEADER -----------------
        ws["A1"] = "Port Performance Report"
        ws["A1"].font = Font(size=14, bold=True)

        ws["A3"] = f"Target: {server}:{port}"
        ws["A4"] = (
            f"Time Range: "
            f"{start_ist.strftime('%Y-%m-%d %H:%M:%S IST')} → "
            f"{end_ist.strftime('%Y-%m-%d %H:%M:%S IST')}"
        )
        ws["A5"] = f"Interval: {interval}"
        ws["A6"] = "(All times shown in IST)"

        for r in range(1, 7):
            ws[f"A{r}"].font = Font(bold=True)

        # ----------------- TABLE HEADERS -----------------
        header_row = 8
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col, value=h)
            c.font = Font(bold=True)
            c.fill = fill_header
            c.alignment = Alignment(horizontal="center")
            c.border = border

        # ----------------- DATA ROWS -----------------
        rows = _query_port_timeseries(server, port, start, end, interval)

        rptr = header_row + 1
        odd = False

        for row in rows:
            # Use IST time but drop tzinfo – Excel does not support tz-aware datetimes
            dt_naive = row["time"].replace(tzinfo=None)

            values = [
                dt_naive,
                round(row["mean_ms"], 2),
                round(row["max_ms"], 2),
                row["attempts"],
                row["failures"],
                round(row["availability"], 2),
            ]

            for col, val in enumerate(values, start=1):
                c = ws.cell(row=rptr, column=col, value=val)
                c.border = border

                if isinstance(val, datetime):
                    c.number_format = "yyyy-mm-dd hh:mm:ss"
                    c.alignment = Alignment(horizontal="center")
                else:
                    c.alignment = Alignment(horizontal="center")

                if odd:
                    c.fill = fill_altrow

            odd = not odd
            rptr += 1

        data_end_row = rptr - 1

        # ----------------- CONDITIONAL FORMATTING -----------------
        if data_end_row >= header_row + 1:
            # Colors
            RED = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            AMBER = PatternFill(start_color="FFDD99", end_color="FFDD99", fill_type="solid")
            LORANGE = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            GREEN = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

            # Column letters
            mean_col_letter = get_column_letter(2)  # "Mean Resp (ms)"
            max_col_letter = get_column_letter(3)   # "Max Resp (ms)"
            avail_col_letter = get_column_letter(6) # "Availability (%)"

            mean_range = f"{mean_col_letter}{header_row+1}:{mean_col_letter}{data_end_row}"
            max_range = f"{max_col_letter}{header_row+1}:{max_col_letter}{data_end_row}"
            avail_range = f"{avail_col_letter}{header_row+1}:{avail_col_letter}{data_end_row}"

            # ------------------------------------
            # RESPONSE TIME (Mean + Max)
            # Rules:
            #   > 1000 → Red
            #   > 500  → Amber
            #   > 250  → Light Orange
            # (Note: rules are overlapping; Excel applies them in order, last rule wins.
            #  We want the strongest to dominate, so we add from lowest to highest.)
            # ------------------------------------

            # Light Orange >250
            ws.conditional_formatting.add(
                mean_range,
                CellIsRule(operator="greaterThan", formula=["250"], fill=LORANGE),
            )
            ws.conditional_formatting.add(
                max_range,
                CellIsRule(operator="greaterThan", formula=["250"], fill=LORANGE),
            )

            # Amber >500 (overrides light orange)
            ws.conditional_formatting.add(
                mean_range,
                CellIsRule(operator="greaterThan", formula=["500"], fill=AMBER),
            )
            ws.conditional_formatting.add(
                max_range,
                CellIsRule(operator="greaterThan", formula=["500"], fill=AMBER),
            )

            # Red >1000 (overrides amber)
            ws.conditional_formatting.add(
                mean_range,
                CellIsRule(operator="greaterThan", formula=["1000"], fill=RED),
            )
            ws.conditional_formatting.add(
                max_range,
                CellIsRule(operator="greaterThan", formula=["1000"], fill=RED),
            )

            # ------------------------------------
            # AVAILABILITY
            # Rules:
            #   < 50  → Red
            #   50–75 → Amber
            #   >=75  → Green
            # We'll use formula-based rules so they don't overlap incorrectly.
            # ------------------------------------
            first_data_row = header_row + 1

            # RED: < 50
            ws.conditional_formatting.add(
                avail_range,
                CellIsRule(
                    operator="lessThan",
                    formula=["50"],
                    fill=RED,
                ),
            )

            # AMBER: >=50 AND <75
            amber_formula = (
                f'AND({avail_col_letter}{first_data_row}>="50",'
                f'{avail_col_letter}{first_data_row}<"75")'
            )
            # However, CellIsRule doesn't support AND; use FormulaRule instead:
            ws.conditional_formatting.add(
                avail_range,
                FormulaRule(
                    formula=[
                        f'AND({avail_col_letter}{first_data_row}>=50,'
                        f'{avail_col_letter}{first_data_row}<75)'
                    ],
                    fill=AMBER,
                ),
            )

            # GREEN: >= 75
            ws.conditional_formatting.add(
                avail_range,
                CellIsRule(
                    operator="greaterThanOrEqual",
                    formula=["75"],
                    fill=GREEN,
                ),
            )

        # ----------------- AUTO COLUMN WIDTHS -----------------
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22

    wb.save(outfile)
    return outfile

