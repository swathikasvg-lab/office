"""
Excel Generator for Server Availability Report (1001)
Matching styling & structure of FINAL generator_pdf.py
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


# ----------------------------
# Excel Builder
# ----------------------------
def build_excel(results, start_ts, end_ts, report_name="Server Availability"):

    output_dir = "/usr/local/autointelli/opsduty-server/generated_reports"
    os.makedirs(output_dir, exist_ok=True)

    excel_path = f"{output_dir}/Server_Availability_Report.xlsx"

    # Remove old file
    try:
        if os.path.exists(excel_path):
            os.remove(excel_path)
    except:
        pass

    wb = openpyxl.Workbook()

    # STYLES
    header_fill = PatternFill("solid", fgColor="1A73E8")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="1A73E8"),
        right=Side(style="thin", color="1A73E8"),
        top=Side(style="thin", color="1A73E8"),
        bottom=Side(style="thin", color="1A73E8"),
    )

    # ============================================================
    # 1) SUMMARY SHEET
    # ============================================================
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "Server Availability Report"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Period: {start_ts.strftime('%Y-%m-%d %H:%M')}  to  {end_ts.strftime('%Y-%m-%d %H:%M')}"
    ws.merge_cells("A2:F2")

    if results:
        ws["A3"] = f"Customer: {results[0]['customer']}"
    else:
        ws["A3"] = "Customer: All Customers"
    ws.merge_cells("A3:F3")

    ws.append([])

    total_servers = len(results)
    avg_availability = (
        sum([float(r.get("availability") or 0) for r in results]) / total_servers
        if total_servers else 0
    )

    # Convert downtime ("1 day 2 hrs 5 mins") to minutes if possible
    def parse_minutes(s):
        import re
        if not s:
            return 0
        s = s.lower()
        mins = 0
        d = re.search(r"(\d+)\s*day", s)
        h = re.search(r"(\d+)\s*hr", s)
        m = re.search(r"(\d+)\s*min", s)
        if d: mins += int(d.group(1)) * 1440
        if h: mins += int(h.group(1)) * 60
        if m: mins += int(m.group(1))
        return mins

    avg_downtime = (
        sum([parse_minutes(r.get("downtime")) for r in results]) / total_servers
        if total_servers else 0
    )

    ws.append(["Total Servers", total_servers])
    ws.append(["Average Availability (%)", f"{avg_availability:.2f}"])
    ws.append(["Average Downtime (mins)", f"{avg_downtime:.2f}"])
    ws.append([])

    # ------------------------------------------------------------
    # Pie Chart: Availability vs Downtime
    # ------------------------------------------------------------
    ws.append(["Metric", "Value"])
    ws.append(["Available %", avg_availability])
    ws.append(["Downtime %", round(100 - avg_availability, 2)])

    chart = PieChart()
    chart.title = "Overall Availability"
    data = Reference(ws, min_col=2, min_row=ws.max_row - 1, max_row=ws.max_row)
    labels = Reference(ws, min_col=1, min_row=ws.max_row - 1, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)

    ws.add_chart(chart, "D6")

    ws.append([])
    ws.append(["Top 10 Downtime Servers"])
    ws.append(["Instance", "Downtime (mins)"])

    # Sort by downtime DESC
    sorted_dt = sorted(
        results,
        key=lambda r: parse_minutes(r.get("downtime")),
        reverse=True
    )[:10]

    for r in sorted_dt:
        ws.append([
            r["instance"],
            parse_minutes(r.get("downtime"))
        ])

    # ============================================================
    # 2) DETAILED DATA SHEET
    # ============================================================
    ws2 = wb.create_sheet(title="Details")
    headers = ["Instance", "Customer", "Availability (%)", "Downtime", "Downtime (mins)"]
    ws2.append(headers)

    # Apply header style
    for c in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for r in results:
        mins = parse_minutes(r.get("downtime"))
        ws2.append([
            r.get("instance"),
            r.get("customer"),
            float(r.get("availability") or 0),
            r.get("downtime"),
            mins
        ])

    # Alternate row background
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="F2F2F2")

    # Auto column size
    for col in ws2.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # ============================================================
    # SAVE
    # ============================================================
    wb.save(excel_path)
    return excel_path

