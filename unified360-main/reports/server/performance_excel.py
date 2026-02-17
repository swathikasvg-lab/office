import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

def build_excel(results, start_ts, end_ts):
    """
    results: same structure as PDF generator
    """
    output_dir = "/usr/local/autointelli/opsduty-server/generated_reports"
    os.makedirs(output_dir, exist_ok=True)
    file_path = f"{output_dir}/Server_Performance_Report.xlsx"
    if os.path.exists(file_path):
        os.remove(file_path)

    wb = openpyxl.Workbook()

    # --- Summary sheet ---
    sum_ws = wb.active
    sum_ws.title = "Summary"

    sum_ws["A1"] = "Server Performance Report"
    sum_ws["A1"].font = Font(size=16, bold=True)
    sum_ws.merge_cells("A1:H1")

    sum_ws["A2"] = f"Period: {start_ts.strftime('%Y-%m-%d %H:%M')} to {end_ts.strftime('%Y-%m-%d %H:%M')}"
    sum_ws.merge_cells("A2:H2")

    sum_ws.append([])
    total_servers = len(results)
    avg_cpu = round(sum([r.get("cpu", 0) for r in results]) / total_servers, 2) if total_servers else 0
    avg_mem = round(sum([r.get("mem", 0) for r in results]) / total_servers, 2) if total_servers else 0

    sum_ws.append(["Total Servers", total_servers])
    sum_ws.append(["Average CPU (%)", avg_cpu])
    sum_ws.append(["Average Memory (%)", avg_mem])
    sum_ws.append([])

    # Top CPU table
    top_cpu = sorted(results, key=lambda x: x.get("cpu", 0), reverse=True)[:10]
    sum_ws.append(["Top CPU Servers"])
    sum_ws.append(["Instance", "CPU (%)"])
    for r in top_cpu:
        sum_ws.append([r["instance"], r["cpu"]])

    # Create chart for Top CPU
    chart = BarChart()
    chart.title = "Top CPU Servers"
    chart.y_axis.title = "CPU (%)"
    chart.x_axis.title = "Instance"

    data_ref = Reference(sum_ws, min_col=2, min_row= sum_ws.max_row - len(top_cpu) + 1, max_row=sum_ws.max_row)
    cats_ref = Reference(sum_ws, min_col=1, min_row= sum_ws.max_row - len(top_cpu) + 1, max_row=sum_ws.max_row)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.shape = 4
    sum_ws.add_chart(chart, "J2")

    # --- Data sheet ---
    data_ws = wb.create_sheet(title="Data")
    headers = ["Instance", "Customer", "CPU (%)", "Memory (%)", "Disk Summary", "Disk Read (KB/s)", "Disk Write (KB/s)", "Net Rx (KB/s)", "Net Tx (KB/s)"]
    data_ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1A73E8")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(headers, 1):
        c = data_ws.cell(row=4, column=col_idx)  # will move header to row 4 to keep space
    # write actual header row at row 1 (already appended)
    for col_idx, h in enumerate(headers, 1):
        cell = data_ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # append rows
    for r in results:
        disk_summary = r.get("disk") or ""
        if isinstance(disk_summary, str) and "," in disk_summary:
            # keep combined but also newline-friendly
            disk_summary = "\n".join([p.strip() for p in disk_summary.split(",")])
        data_ws.append([
            r.get("instance"),
            r.get("customer", ""),
            r.get("cpu", 0),
            r.get("mem", 0),
            disk_summary,
            r.get("disk_read", 0),
            r.get("disk_write", 0),
            r.get("net_in", 0),
            r.get("net_out", 0)
        ])

    # Auto-width columns
    for col in data_ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        data_ws.column_dimensions[col_letter].width = min(60, max_len + 2)

    # Add footer meta in Summary
    sum_ws.append([])
    sum_ws.append([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])

    wb.save(file_path)
    return file_path

