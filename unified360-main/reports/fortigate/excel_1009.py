# reports/fortigate/excel_1009.py
import os
import tempfile
from datetime import datetime, timezone, timedelta
import requests
from flask import current_app
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

IST = timezone(timedelta(hours=5, minutes=30))

# styles
thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
hdr_fill = PatternFill("solid", fgColor="F0F4FF")
up_fill = PatternFill("solid", fgColor="CCFFCC")
down_fill = PatternFill("solid", fgColor="FFCCCC")
amber_fill = PatternFill("solid", fgColor="FFB84D")
light_orange = PatternFill("solid", fgColor="FFE6CC")

def _safe_float(x):
    try:
        if x is None: return 0.0
        return float(x)
    except:
        return 0.0

def human_bytes_auto(b):
    if b is None:
        return "0 B"
    try:
        b = float(b)
    except:
        return "0 B"
    if b < 1024: return f"{int(b)} B"
    kb = b/1024.0
    if kb < 1024: return f"{kb:.2f} KB"
    mb = kb/1024.0
    if mb < 1024: return f"{mb:.2f} MB"
    gb = mb/1024.0
    if gb < 1024: return f"{gb:.2f} GB"
    tb = gb/1024.0
    return f"{tb:.2f} TB"

# reuse same fetch as PDF (but simpler calls)
def _fetch_links(device, start, end):
    influx_url = current_app.config["INFLUXDB_URL"]
    influx_db  = current_app.config["INFLUXDB_DB"]
    q = f"""
    SELECT LAST("fgVWLHealthCheckLinkLatency") AS latency,
           LAST("fgVWLHealthCheckLinkJitter") AS jitter,
           LAST("fgVWLHealthCheckLinkPacketLoss") AS packet_loss,
           LAST("fgVWLHealthCheckLinkState") AS state,
           LAST("fgVWLHealthCheckLinkBandwidthIn") AS bw_in,
           LAST("fgVWLHealthCheckLinkBandwidthOut") AS bw_out,
           LAST("fgVWLHealthCheckLinkUsedBandwidthIn") AS used_in,
           LAST("fgVWLHealthCheckLinkUsedBandwidthOut") AS used_out,
           LAST("fgVWLHealthCheckLinkIfName") AS iface,
           LAST("fgVWLHealthCheckLinkMOS") AS mos,
           LAST("fgVWLHealthCheckLinkName") AS linkname
    FROM "sdwan_health"
    WHERE "hostname" = '{device}'
      AND time >= '{start}' AND time <= '{end}'
    GROUP BY "hc_name"
    """
    resp = requests.get(influx_url, params={"db": influx_db, "q": q}, timeout=30).json()
    series = resp.get("results", [{}])[0].get("series", [])
    links = []
    for s in series:
        tags = s.get("tags", {})
        hc = tags.get("hc_name") or "unknown"
        v = s.get("values", [])
        if not v:
            continue
        v = v[-1]
        links.append({
            "name": v[10] or hc,
            "latency": _safe_float(v[1]),
            "jitter": _safe_float(v[2]),
            "packet_loss": _safe_float(v[3]),
            "state": int(v[4] or 0),
            "bw_in": int(v[5] or 0),
            "bw_out": int(v[6] or 0),
            "used_in": int(v[7] or 0),
            "used_out": int(v[8] or 0),
            "iface": v[9] or "",
            "mos": v[10] or ""
        })
    return links

def _write_table(ws, start_row, headers, rows, status_col_idx=None, color_loss=False):
    # headers
    r = start_row
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
    r += 1
    for row in rows:
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = thin
            # status coloring
            if status_col_idx and c-1 == status_col_idx:
                if v == "UP":
                    cell.fill = up_fill
                elif v == "DOWN":
                    cell.fill = down_fill
            # packet loss coloring
            if color_loss and (len(row) > 3):
                try:
                    pct = float(row[3])
                except:
                    pct = 0.0
                if pct >= 25:
                    ws.cell(row=r, column=4).fill = down_fill
                elif pct >= 10:
                    ws.cell(row=r, column=4).fill = amber_fill
                elif pct >=5:
                    ws.cell(row=r, column=4).fill = light_orange
        r += 1
    # auto width
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    return r+1

def build_excel(device, start, end, interval):
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    outfile = os.path.join(tempfile.gettempdir(), f"rpt_1009_sdwan_{device}_{ts}.xlsx")
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    links = _fetch_links(device, start, end)
    total = len(links)
    down_count = len([l for l in links if l["state"] != 0])

    # Summary sheet
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Fortigate SD-WAN Performance Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Device"; ws["B3"] = device
    ws["A4"] = "From"; ws["B4"] = datetime.fromisoformat(start.replace("Z","+00:00")).astimezone(IST).strftime("%Y-%m-%d %H:%M:%S")
    ws["A5"] = "To"; ws["B5"] = datetime.fromisoformat(end.replace("Z","+00:00")).astimezone(IST).strftime("%Y-%m-%d %H:%M:%S")
    ws["A7"] = "Total Links"; ws["B7"] = total
    ws["A8"] = "Down Links"; ws["B8"] = down_count

    # Top 10 latency
    ws_lat = wb.create_sheet("Top_Latency")
    top_lat = sorted(links, key=lambda x: -x["latency"])[:10]
    rows = []
    for l in top_lat:
        rows.append([l["name"], f"{l['latency']:.3f}", f"{l['jitter']:.3f}", f"{l['packet_loss']:.3f}", ("UP" if l["state"]==0 else "DOWN")])
    _write_table(ws_lat, 1, ["Link","Latency(ms)","Jitter(ms)","PacketLoss(%)","State"], rows, status_col_idx=4, color_loss=True)

    # Top 10 jitter
    ws_jit = wb.create_sheet("Top_Jitter")
    top_jit = sorted(links, key=lambda x: -x["jitter"])[:10]
    rows=[]
    for l in top_jit:
        rows.append([l["name"], f"{l['jitter']:.3f}", f"{l['latency']:.3f}", f"{l['packet_loss']:.3f}", ("UP" if l["state"]==0 else "DOWN")])
    _write_table(ws_jit,1,["Link","Jitter(ms)","Latency(ms)","PacketLoss(%)","State"],rows,status_col_idx=4,color_loss=True)

    # Down links
    ws_down = wb.create_sheet("Down_Links")
    down = [l for l in links if l["state"]!=0]
    rows=[]
    for l in down:
        rows.append([l["name"], l["iface"], f"{l['latency']:.3f}", f"{l['jitter']:.3f}", f"{l['packet_loss']:.3f}", "DOWN"])
    _write_table(ws_down,1,["Link","Interface","Latency(ms)","Jitter(ms)","PacketLoss(%)","State"],rows,status_col_idx=5,color_loss=True)

    # Top packet loss
    ws_loss = wb.create_sheet("Top_PacketLoss")
    top_loss = sorted(links, key=lambda x: -x["packet_loss"])[:10]
    rows=[]
    for l in top_loss:
        rows.append([l["name"], f"{l['packet_loss']:.3f}", f"{l['latency']:.3f}", f"{l['jitter']:.3f}", ("UP" if l["state"]==0 else "DOWN")])
    _write_table(ws_loss,1,["Link","PacketLoss(%)","Latency(ms)","Jitter(ms)","State"],rows,status_col_idx=4,color_loss=True)

    # Detailed
    ws_det = wb.create_sheet("Detailed")
    headers = ["Link","Interface","Latency(ms)","Jitter(ms)","PacketLoss(%)","State","BW In","BW Out","Used In","Used Out","MOS"]
    det_rows=[]
    for l in links:
        det_rows.append([l["name"], l["iface"], f"{l['latency']:.3f}", f"{l['jitter']:.3f}", f"{l['packet_loss']:.3f}", ("UP" if l["state"]==0 else "DOWN"), human_bytes_auto(l["bw_in"]), human_bytes_auto(l["bw_out"]), human_bytes_auto(l["used_in"]), human_bytes_auto(l["used_out"]), l["mos"]])
    _write_table(ws_det,1,headers,det_rows,status_col_idx=5,color_loss=True)

    wb.save(outfile)
    return outfile

