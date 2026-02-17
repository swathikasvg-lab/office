# reports/fortigate/excel_1008.py
import os
import tempfile
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .pdf_1008 import IST, _fetch_tunnels_latest, _fetch_volume_deltas, _secs_to_rounded_human, _human_bytes_auto

def build_excel(device, start, end, interval):
    out_dir = tempfile.gettempdir()
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    outfile = os.path.join(out_dir, f"rpt_1008_fortigate_{device}_{ts}.xlsx")

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    tunnels = _fetch_tunnels_latest(device, start, end)
    deltas = _fetch_volume_deltas(device, start, end)

    for t in tunnels:
        vpn = t["vpn_name"]
        d = deltas.get(vpn, {"in_delta": 0, "out_delta": 0})
        t["in_delta"] = d["in_delta"]
        t["out_delta"] = d["out_delta"]

    total = len(tunnels)
    up = sum(1 for t in tunnels if t.get("status_code") == 2)
    down = sum(1 for t in tunnels if t.get("status_code") == 1)

    # Summary sheet
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Fortigate VPN Performance Report"
    ws["A1"].font = Font(size=14, bold=True)
    start_ist = datetime.fromisoformat(start.replace("Z", "+00:00")).astimezone(IST)
    end_ist = datetime.fromisoformat(end.replace("Z", "+00:00")).astimezone(IST)
    ws["A3"] = f"Device: {device}"
    ws["A4"] = f"Range: {start_ist.strftime('%Y-%m-%d %H:%M:%S IST')} → {end_ist.strftime('%Y-%m-%d %H:%M:%S IST')}"
    ws["A5"] = f"Interval: {interval}"
    ws["A7"] = "Total Tunnels"; ws["B7"] = total
    ws["A8"] = "UP"; ws["B8"] = up
    ws["A9"] = "DOWN"; ws["B9"] = down
    for c in ["A7","A8","A9"]:
        ws[c].font = Font(bold=True)

    # Top 10 by In
    top_in = sorted(tunnels, key=lambda x: -x.get("in_delta", 0))[:10]
    ws["A11"] = "Top 10 by In Volume"
    row = 12
    ws.cell(row=row, column=1, value="VPN Name"); ws.cell(row=row, column=2, value="In Volume"); ws.cell(row=row, column=3, value="Out Volume")
    for c in range(1,4): ws.cell(row=row, column=c).font = Font(bold=True)
    row += 1
    for t in top_in:
        ws.cell(row=row, column=1, value=t["vpn_name"])
        ws.cell(row=row, column=2, value=_human_bytes_auto(t.get("in_delta", 0)))
        ws.cell(row=row, column=3, value=_human_bytes_auto(t.get("out_delta", 0)))
        row += 1

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 28

    # Detailed sheet
    ws2 = wb.create_sheet("Detailed")
    headers = ["VPN Name","Status","Phase2","Remote GW","Uptime","In Volume","Out Volume"]
    for c, h in enumerate(headers, start=1):
        ws2.cell(row=1, column=c, value=h).font = Font(bold=True)
        ws2.column_dimensions[get_column_letter(c)].width = 30
    r = 2
    sorted_tunnels = sorted(tunnels, key=lambda x: ((0 if x.get("status_code")==2 else 1), -x.get("in_delta",0)))
    for t in sorted_tunnels:
        ws2.cell(row=r, column=1, value=t["vpn_name"])
        st = "UP" if t.get("status_code")==2 else ("DOWN" if t.get("status_code")==1 else "UNKNOWN")
        stc = ws2.cell(row=r, column=2, value=st)
        ws2.cell(row=r, column=3, value=t.get("phase2",""))
        ws2.cell(row=r, column=4, value=t.get("remote_gw",""))
        ws2.cell(row=r, column=5, value=_secs_to_rounded_human(t.get("life_secs",0)))
        ws2.cell(row=r, column=6, value=_human_bytes_auto(t.get("in_delta",0)))
        ws2.cell(row=r, column=7, value=_human_bytes_auto(t.get("out_delta",0)))
        # color status
        if t.get("status_code")==2:
            stc.fill = PatternFill("solid", fgColor="CCFFCC")
        elif t.get("status_code")==1:
            stc.fill = PatternFill("solid", fgColor="FFCCCC")
        else:
            stc.fill = PatternFill("solid", fgColor="EDEDED")
        r += 1

    # Raw traffic deltas
    ws3 = wb.create_sheet("TrafficDeltas")
    ws3.append(["VPN Name", "In Delta (bytes)", "Out Delta (bytes)"])
    for t in sorted_tunnels:
        ws3.append([t["vpn_name"], t.get("in_delta",0), t.get("out_delta",0)])
    for col in range(1,4):
        ws3.column_dimensions[get_column_letter(col)].width = 26

    wb.save(outfile)
    return outfile

