# reports/url/excel_1004.py

from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .url_data import UrlDataFetcher


class ExcelUrlReport:
    """
    Generates URL Performance Report (Excel) for one or more URLs.
    """

    def __init__(self):
        self.data_fetcher = UrlDataFetcher()

    def _autosize_columns(self, ws):
        for column_cells in ws.columns:
            length = 0
            col = column_cells[0].column
            for cell in column_cells:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                except Exception:
                    cell_len = 0
                if cell_len > length:
                    length = cell_len
            ws.column_dimensions[get_column_letter(col)].width = length + 2

    def generate(self, urls, start: str, end: str) -> str:
        """
        urls: list of server strings
        start, end: datetime strings from UI
        Returns: path to generated XLSX file
        """
        if isinstance(urls, str):
            urls = [urls]

        ts = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        outfile = f"/tmp/url_performance_{ts}.xlsx"

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        ws_summary.append(["URL Performance Report"])
        ws_summary.append([f"Report Period: {start} to {end}"])
        ws_summary.append([f"Generated On: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"])
        ws_summary.append([])

        header = [
            "Server (URL)",
            "Friendly Name",
            "Total Checks",
            "Success Checks",
            "Failed Checks",
            "Availability (%)",
            "Avg Response Time (s)",
            "Min Response Time (s)",
            "Max Response Time (s)",
            "SSL Common Name",
            "SSL Issuer",
            "SSL Expiry Date",
            "SSL Days Left",
            "SSL Verification",
        ]
        ws_summary.append(header)

        # Status codes sheet
        ws_codes = wb.create_sheet(title="Status Codes")
        ws_codes.append(["Server (URL)", "Status Code", "Count"])

        for server in urls:
            summary = self.data_fetcher.get_summary(server, start, end)

            ssl_info = self.data_fetcher.get_ssl_info(summary["friendly_name"])
            ssl_common_name = ssl_info["common_name"] if ssl_info and ssl_info.get("common_name") else None
            ssl_issuer = ssl_info["issuer"] if ssl_info and ssl_info.get("issuer") else None
            ssl_expiry = ssl_info["expiry_date"] if ssl_info and ssl_info.get("expiry_date") else None
            ssl_days_left = ssl_info["days_left"] if ssl_info and ssl_info.get("days_left") is not None else None
            ssl_verification = ssl_info["verification"] if ssl_info and ssl_info.get("verification") else None

            ws_summary.append(
                [
                    summary["server"],
                    summary["friendly_name"],
                    summary["total_checks"],
                    summary["success_checks"],
                    summary["failed_checks"],
                    summary["availability_pct"],
                    summary["avg_response_time"],
                    summary["min_response_time"],
                    summary["max_response_time"],
                    ssl_common_name,
                    ssl_issuer,
                    ssl_expiry,
                    ssl_days_left,
                    ssl_verification,
                ]
            )

            # Status codes
            codes = self.data_fetcher.get_status_codes(server, start, end)
            if not codes:
                continue
            for item in codes:
                ws_codes.append(
                    [
                        summary["server"],
                        item["status_code"],
                        item["count"],
                    ]
                )

        self._autosize_columns(ws_summary)
        self._autosize_columns(ws_codes)

        wb.save(outfile)
        return outfile

