import os
import tempfile
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from .pdf_1006 import IST, _query_ping_timeseries


def build_excel(urls, start, end, interval):
    """
    Build Excel Ping Performance Report with:
      - one sheet per target
      - IST timestamps
      - latency & packet loss conditional formatting
    """
    out_dir = tempfile.gettempdir()
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    outfile = os.path.join(out_dir, f"rpt_1006_ping_{ts}.xlsx")

    wb = openpyxl.Workbook()

    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_header = PatternFill("solid", fgColor="DCE6F7")
    fill_altrow = PatternFill("solid", fgColor="F7FAFF")

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    start_ist = datetime.fromisoformat(start.replace("Z", "+00:00")).astimezone(IST)
    end_ist = datetime.fromisoformat(end.replace("Z", "+00:00")).astimezone(IST)

    headers = [
        "Time (IST)",
        "Avg RTT (ms)",
        "Min RTT (ms)",
        "Max RTT (ms)",
        "Std Dev (ms)",
        "Packets Tx",
        "Packets Rx",
        "Loss (%)",
    ]

    for idx, url in enumerate(urls):
        sheet_name = url[:31] or f"target_{idx+1}"
        ws = wb.create_sheet(title=sheet_name)

        # ----------------- SHEET HEADER -----------------
        ws["A1"] = "Ping Performance Report"
        ws["A1"].font = Font(size=14, bold=True)

        ws["A3"] = f"Target: {url}"
        ws["A4"] = (
            f"Time Range: "
            f"{start_ist.strftime('%Y-%m-%d %H:%M:%S IST')} → "
            f"{end_ist.strftime('%Y-%m-%d %H:%M:%S IST')}"
        )
        ws["A5"] = f"Interval: {interval}"
        ws["A6"] = "(All times shown in IST)"

        for r in range(1, 7):
            ws[f"A{r}"].font = Font(bold=True)

        # ----------------- TABLE HEADERS -----------------
        header_row = 8
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col, value=h)
            c.font = Font(bold=True)
            c.fill = fill_header
            c.alignment = Alignment(horizontal="center")
            c.border = border

        # ----------------- DATA ROWS -----------------
        rows = _query_ping_timeseries(url, start, end, interval)

        rptr = header_row + 1
        odd = False

        for row in rows:
            dt_naive = row["time"].replace(tzinfo=None)

            values = [
                dt_naive,
                round(row["avg_ms"], 2),
                round(row["min_ms"], 2),
                round(row["max_ms"], 2),
                round(row["stddev_ms"], 2),
                row["tx"],
                row["rx"],
                round(row["loss_pct"], 2),
            ]

            for col, val in enumerate(values, start=1):
                c = ws.cell(row=rptr, column=col, value=val)
                c.border = border

                if isinstance(val, datetime):
                    c.number_format = "yyyy-mm-dd hh:mm:ss"
                    c.alignment = Alignment(horizontal="center")
                else:
                    c.alignment = Alignment(horizontal="center")

                if odd:
                    c.fill = fill_altrow

            odd = not odd
            rptr += 1

        data_end_row = rptr - 1

        # ----------------- CONDITIONAL FORMATTING -----------------
        if data_end_row >= header_row + 1:
            # Latency columns: Avg (2), Min (3), Max (4), StdDev (5)
            latency_cols = [2, 3, 4, 5]

            light_orange = "FFF7D5"  # >250 ms
            amber = "FFBF00"        # >500 ms
            red = "FF4C4C"          # >1000 ms

            for col in latency_cols:
                col_letter = get_column_letter(col)
                data_range = f"{col_letter}{header_row+1}:{col_letter}{data_end_row}"

                # >250 ms – light orange
                ws.conditional_formatting.add(
                    data_range,
                    CellIsRule(
                        operator="greaterThan",
                        formula=["250"],
                        fill=PatternFill(
                            start_color=light_orange,
                            end_color=light_orange,
                            fill_type="solid",
                        ),
                    ),
                )

                # >500 ms – amber
                ws.conditional_formatting.add(
                    data_range,
                    CellIsRule(
                        operator="greaterThan",
                        formula=["500"],
                        fill=PatternFill(
                            start_color=amber,
                            end_color=amber,
                            fill_type="solid",
                        ),
                    ),
                )

                # >1000 ms – red
                ws.conditional_formatting.add(
                    data_range,
                    CellIsRule(
                        operator="greaterThan",
                        formula=["1000"],
                        fill=PatternFill(
                            start_color=red,
                            end_color=red,
                            fill_type="solid",
                        ),
                    ),
                )

            # Packet loss column: Loss (%) = 8
            loss_col_letter = get_column_letter(8)
            loss_range = f"{loss_col_letter}{header_row+1}:{loss_col_letter}{data_end_row}"

            # ≥25% – light yellow
            ws.conditional_formatting.add(
                loss_range,
                CellIsRule(
                    operator="greaterThanOrEqual",
                    formula=["25"],
                    fill=PatternFill(
                        start_color="FFF8CC",
                        end_color="FFF8CC",
                        fill_type="solid",
                    ),
                ),
            )

            # ≥50% – amber
            ws.conditional_formatting.add(
                loss_range,
                CellIsRule(
                    operator="greaterThanOrEqual",
                    formula=["50"],
                    fill=PatternFill(
                        start_color="FFD966",
                        end_color="FFD966",
                        fill_type="solid",
                    ),
                ),
            )

            # ≥75% – red
            ws.conditional_formatting.add(
                loss_range,
                CellIsRule(
                    operator="greaterThanOrEqual",
                    formula=["75"],
                    fill=PatternFill(
                        start_color="FF4C4C",
                        end_color="FF4C4C",
                        fill_type="solid",
                    ),
                ),
            )

        # ----------------- AUTO COLUMN WIDTHS -----------------
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22

    wb.save(outfile)
    return outfile

