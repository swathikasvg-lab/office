import os
import tempfile
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from flask import current_app

# Reuse from PDF module
from .pdf_1005 import (
    IST,
    _INTERVAL_SECONDS,
    _format_bytes,
    _format_speed_kbps,
)
from .pdf_1005 import _query_interface_timeseries


def build_excel(template_type, device, interfaces, start, end, interval):

    out_dir = tempfile.gettempdir()
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    outfile = os.path.join(out_dir, f"rpt_1005_bandwidth_{ts}.xlsx")

    wb = openpyxl.Workbook()

    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_header = PatternFill("solid", fgColor="DCE6F7")
    fill_altrow = PatternFill("solid", fgColor="F7FAFF")

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # =====================================================
    #              PER-INTERFACE SHEETS
    # =====================================================
    for idx, iface in enumerate(interfaces):
        sheet_name = iface[:31] or f"iface_{idx+1}"
        ws = wb.create_sheet(title=sheet_name)

        # --------------------- HEADER ----------------------
        ws["A1"] = "Bandwidth Utilization Report"
        ws["A1"].font = Font(size=14, bold=True)

        ws["A3"] = f"Template Type: {template_type or '-'}"
        ws["A4"] = f"Device: {device}"
        ws["A5"] = f"Interface: {iface}"

        # Convert selected time to IST
        start_ist = datetime.fromisoformat(start.replace("Z", "+00:00")).astimezone(IST)
        end_ist   = datetime.fromisoformat(end.replace("Z", "+00:00")).astimezone(IST)

        ws["A6"] = f"Time Range: {start_ist.strftime('%Y-%m-%d %H:%M:%S IST')} → {end_ist.strftime('%Y-%m-%d %H:%M:%S IST')}"
        ws["A7"] = f"Interval: {interval}"
        ws["A8"] = "(All times shown in IST)"

        for r in range(1, 9):
            ws[f"A{r}"].font = Font(bold=True)

        # ------------------- TABLE HEADERS -------------------
        headers = [
            "Time (IST)",
            "Traffic Total (Volume)",
            "Traffic Total (Speed)",
            "Traffic In (Volume)",
            "Traffic In (Speed)",
            "Traffic Out (Volume)",
            "Traffic Out (Speed)",
        ]

        header_row = 10

        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col, value=h)
            c.font = Font(bold=True)
            c.fill = fill_header
            c.alignment = Alignment(horizontal="center")
            c.border = border

        # -------------------- FETCH DATA ---------------------
        rows = _query_interface_timeseries(device, iface, start, end, interval)

        row_ptr = header_row + 1
        odd = False

        for row in rows:
            excel_dt = row["time"].replace(tzinfo=None) 

            values = [
                excel_dt,
                _format_bytes(row["total_bytes"]),
                _format_speed_kbps(row["total_kbps"]),
                _format_bytes(row["in_bytes"]),
                _format_speed_kbps(row["in_kbps"]),
                _format_bytes(row["out_bytes"]),
                _format_speed_kbps(row["out_kbps"]),
            ]

            for col, val in enumerate(values, start=1):
                c = ws.cell(row=row_ptr, column=col, value=val)
                c.border = border

                if isinstance(val, datetime):
                    c.number_format = "yyyy-mm-dd hh:mm:ss"

                if odd:
                    c.fill = fill_altrow

                # align text / numbers
                if col == 1:
                    c.alignment = Alignment(horizontal="center")
                else:
                    c.alignment = Alignment(horizontal="center")

            odd = not odd
            row_ptr += 1

        data_end_row = row_ptr - 1

        # Auto column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 28


    wb.save(outfile)
    return outfile

